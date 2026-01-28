from flask import Blueprint, make_response, request
from flask_login import login_required
from io import BytesIO, StringIO
import csv
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Equipamento
from . import db

exports_bp = Blueprint("exports", __name__)

@exports_bp.route("/export/csv")
@login_required
def export_csv():
    output = StringIO()
    writer = csv.writer(output)
    
    # Cabeçalho
    writer.writerow([
        'ID', 'Nome', 'Categoria', 'Marca', 'Setor', 'Cargo', 
        'Empréstimo', 'Compartilhado', 'AnyDesk', 'Observações', 
        'Criado em', 'Atualizado em'
    ])
    
    # Dados
    for eq in Equipamento.query.yield_per(100):
        writer.writerow([
            eq.id,
            eq.name_response,
            eq.equipamento_category,
            eq.marca_category,
            eq.setor_category or '',
            eq.cargo_category or '',
            eq.emprestimo,
            eq.equipamento_compartilhado,
            eq.numero_anydesk or '',
            eq.observacoes or '',
            eq.created_at.strftime('%d/%m/%Y %H:%M') if eq.created_at else '',
            eq.updated_at.strftime('%d/%m/%Y %H:%M') if eq.updated_at else ''
        ])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=equipamentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return response

@exports_bp.route("/export/excel")
@login_required
def export_excel():
    equipamentos = Equipamento.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    
    # Cabeçalhos
    headers = [
        'ID', 'Nome', 'Categoria', 'Marca', 'Setor', 'Cargo', 
        'Empréstimo', 'Compartilhado', 'AnyDesk', 'Observações', 
        'Criado em', 'Atualizado em'
    ]
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Adicionar dados
    for row, eq in enumerate(equipamentos, 2):
        ws.cell(row=row, column=1, value=eq.id)
        ws.cell(row=row, column=2, value=eq.name_response)
        ws.cell(row=row, column=3, value=eq.equipamento_category)
        ws.cell(row=row, column=4, value=eq.marca_category)
        ws.cell(row=row, column=5, value=eq.setor_category or '')
        ws.cell(row=row, column=6, value=eq.cargo_category or '')
        ws.cell(row=row, column=7, value=eq.emprestimo)
        ws.cell(row=row, column=8, value=eq.equipamento_compartilhado)
        ws.cell(row=row, column=9, value=eq.numero_anydesk or '')
        ws.cell(row=row, column=10, value=eq.observacoes or '')
        ws.cell(row=row, column=11, value=eq.created_at.strftime('%d/%m/%Y %H:%M') if eq.created_at else '')
        ws.cell(row=row, column=12, value=eq.updated_at.strftime('%d/%m/%Y %H:%M') if eq.updated_at else '')
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            cell_value = str(cell.value) if cell.value is not None else ''
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=equipamentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@exports_bp.route("/export/pdf")
@login_required
def export_pdf():
    equipamentos = Equipamento.query.all()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Conteúdo
    story = []
    
    # Título
    title = Paragraph("Relatório de Equipamentos", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Data de geração
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1
    )
    date_text = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", date_style)
    story.append(date_text)
    story.append(Spacer(1, 20))
    
    # Tabela
    data = [['ID', 'Nome', 'Categoria', 'Marca', 'Status', 'Setor']]
    
    for eq in equipamentos:
        status = "Em Uso" if eq.emprestimo == 'SIM' else "Disponível"
        try:
            name = eq.name_response[:25] + '...' if len(eq.name_response) > 25 else eq.name_response
        except (AttributeError, TypeError):
            name = 'N/A'
        try:
            marca = eq.marca_category[:15] + '...' if len(eq.marca_category) > 15 else eq.marca_category
        except (AttributeError, TypeError):
            marca = 'N/A'
        
        data.append([
            str(eq.id),
            name,
            eq.equipamento_category or 'N/A',
            marca,
            status,
            eq.setor_category or 'N/A'
        ])
    
    table = Table(data, colWidths=[0.8*inch, 2.2*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Rodapé com estatísticas
    story.append(Spacer(1, 20))
    total = len(equipamentos)
    em_uso = len([eq for eq in equipamentos if eq.emprestimo == 'SIM'])
    disponiveis = total - em_uso
    
    stats_text = f"Total de equipamentos: {total} | Em uso: {em_uso} | Disponíveis: {disponiveis}"
    stats_para = Paragraph(stats_text, styles['Normal'])
    story.append(stats_para)
    
    doc.build(story)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=equipamentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response